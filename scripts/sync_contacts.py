#!/usr/bin/env python3
"""Validate the private contact workbook and replace Supabase contact rows."""

from __future__ import annotations

import argparse
import json
import os
import re
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


def normalized_name(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_phone(value: Any) -> tuple[str, str]:
    display = normalized_name(value)
    if not display:
        raise ValueError("Hiányzó telefonszám.")
    compact = re.sub(r"[\s()\-/.]", "", display)
    if compact.startswith("00"):
        compact = f"+{compact[2:]}"
    elif compact.startswith("06"):
        compact = f"+36{compact[2:]}"
    elif compact.startswith("36"):
        compact = f"+{compact}"
    if not re.fullmatch(r"\+[1-9][0-9]{7,14}", compact):
        raise ValueError(
            f"Érvénytelen telefonszám: {display}. Használj például +36301234567 formátumot."
        )
    return compact, display


def find_contacts_table(workbook):
    for worksheet in workbook.worksheets:
        for table in worksheet.tables.values():
            if table.name.casefold() == "contacts":
                return worksheet, table
    raise ValueError("Nem található a 'contacts' nevű Excel-tábla.")


def parse_contacts(workbook_path: Path, schedule_path: Path) -> list[dict[str, Any]]:
    schedule = json.loads(schedule_path.read_text(encoding="utf-8"))
    workers_by_name = {
        normalized_name(worker["name"]).casefold(): worker
        for worker in schedule.get("workers", [])
    }
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    worksheet, table = find_contacts_table(workbook)
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = {
        normalized_name(worksheet.cell(min_row, column).value).casefold(): column
        for column in range(min_col, max_col + 1)
    }
    for required in ("név", "telefonszám"):
        if required not in headers:
            raise ValueError(f"Hiányzó kötelező oszlop: {required.title()}")

    contacts: list[dict[str, Any]] = []
    seen: set[str] = set()
    errors: list[str] = []
    for row in range(min_row + 1, max_row + 1):
        name = normalized_name(worksheet.cell(row, headers["név"]).value)
        phone_value = worksheet.cell(row, headers["telefonszám"]).value
        if not name and (phone_value is None or not str(phone_value).strip()):
            continue
        if phone_value is None or not str(phone_value).strip():
            continue
        worker = workers_by_name.get(name.casefold())
        if not worker:
            errors.append(f"{row}. sor: a név nem szerepel az aktív beosztásban: {name or '(hiányzik)'}")
            continue
        if worker["id"] in seen:
            errors.append(f"{row}. sor: duplikált dolgozó: {worker['name']}")
            continue
        try:
            phone_e164, phone_display = normalize_phone(phone_value)
        except ValueError as error:
            errors.append(f"{row}. sor ({worker['name']}): {error}")
            continue
        note_column = headers.get("megjegyzés")
        note = normalized_name(worksheet.cell(row, note_column).value) if note_column else ""
        contacts.append(
            {
                "worker_id": worker["id"],
                "name": worker["name"],
                "phone_e164": phone_e164,
                "phone_display": phone_display,
                "note": note[:500],
            }
        )
        seen.add(worker["id"])

    if errors:
        raise ValueError("\n".join(errors))
    if not contacts:
        raise ValueError("A kontakt Excel egyetlen kitöltött, érvényes telefonszámot sem tartalmaz.")
    return contacts


def rest_request(base_url: str, service_key: str, method: str, path: str, body=None):
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
    }
    if method == "POST":
        headers["Prefer"] = "resolution=merge-duplicates,return=minimal"
    data = json.dumps(body, ensure_ascii=False).encode("utf-8") if body is not None else None
    request = urllib.request.Request(f"{base_url.rstrip('/')}{path}", data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            content = response.read()
            return json.loads(content) if content else None
    except urllib.error.HTTPError as error:
        detail = error.read().decode("utf-8", "replace")
        raise RuntimeError(f"Supabase API-hiba ({error.code}): {detail[:500]}") from error


def sync_contacts(contacts: list[dict[str, Any]], base_url: str, service_key: str) -> None:
    existing = rest_request(
        base_url,
        service_key,
        "GET",
        "/rest/v1/worker_contacts?select=worker_id",
    ) or []
    rest_request(
        base_url,
        service_key,
        "POST",
        "/rest/v1/worker_contacts?on_conflict=worker_id",
        contacts,
    )
    incoming_ids = {contact["worker_id"] for contact in contacts}
    for row in existing:
        worker_id = row.get("worker_id")
        if worker_id and worker_id not in incoming_ids:
            encoded = urllib.parse.quote(worker_id, safe="")
            rest_request(
                base_url,
                service_key,
                "DELETE",
                f"/rest/v1/worker_contacts?worker_id=eq.{encoded}",
            )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=Path("data/Sirok 2026 kontaktok.xlsx"))
    parser.add_argument("--schedule", type=Path, default=Path("site/data/schedule.json"))
    parser.add_argument("--validate-only", action="store_true")
    args = parser.parse_args()

    contacts = parse_contacts(args.workbook, args.schedule)
    if args.validate_only:
        print(f"Érvényes kontaktlista: {len(contacts)} dolgozó.")
        return 0

    supabase_url = os.environ.get("SUPABASE_URL", "").strip()
    service_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    if not supabase_url or not service_key:
        raise RuntimeError("A SUPABASE_URL vagy a SUPABASE_SERVICE_ROLE_KEY nincs beállítva.")
    sync_contacts(contacts, supabase_url, service_key)
    print(f"Supabase kontaktlista frissítve: {len(contacts)} dolgozó.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
