#!/usr/bin/env python3
"""Build the public schedule JSON and information HTML from repository sources."""

from __future__ import annotations

import argparse
import hashlib
import html
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

try:
    import markdown
except ImportError:  # Small, safe fallback for local inspection without requirements.
    markdown = None


SHIFT_HEADER = re.compile(
    r"^(Sze|Cs|P|Szo)\s*\n\s*(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})\s*$",
    re.IGNORECASE,
)
DAY_NAMES = {"sze": "Sze", "cs": "Cs", "p": "P", "szo": "Szo"}


def is_x(value: Any) -> bool:
    return isinstance(value, str) and value.strip().casefold() == "x"


def slugify(value: str) -> str:
    plain = unicodedata.normalize("NFKD", value)
    plain = "".join(char for char in plain if not unicodedata.combining(char))
    slug = re.sub(r"[^a-z0-9]+", "-", plain.casefold()).strip("-")
    return slug or "elem"


def worker_id(name: str) -> str:
    digest = hashlib.sha256(name.strip().casefold().encode("utf-8")).hexdigest()[:10]
    return f"{slugify(name)[:42]}-{digest}"


def sort_key(value: str) -> tuple[str, str]:
    plain = unicodedata.normalize("NFKD", value.casefold())
    plain = "".join(char for char in plain if not unicodedata.combining(char))
    return plain, value.casefold()


def duration_hours(start_hour: int, start_minute: int, end_hour: int, end_minute: int) -> float:
    start = start_hour * 60 + start_minute
    end = end_hour * 60 + end_minute
    if end <= start:
        end += 24 * 60
    return (end - start) / 60


def find_contact_table(workbook):
    for worksheet in workbook.worksheets:
        for table in worksheet.tables.values():
            if table.name.casefold() == "contact":
                return worksheet, table
    raise ValueError("Nem található a 'contact' nevű Excel-tábla.")


def continuous_blocks(assignments: dict[str, bool], shifts: list[dict[str, Any]]) -> list[dict[str, Any]]:
    blocks: list[dict[str, Any]] = []
    start_index: int | None = None
    for index in range(len(shifts) + 1):
        active = index < len(shifts) and assignments[shifts[index]["id"]]
        if active and start_index is None:
            start_index = index
        if not active and start_index is not None:
            selected = shifts[start_index:index]
            blocks.append(
                {
                    "startShiftId": selected[0]["id"],
                    "endShiftId": selected[-1]["id"],
                    "startLabel": f'{selected[0]["day"]} {selected[0]["start"]}',
                    "endLabel": f'{selected[-1]["day"]} {selected[-1]["end"]}',
                    "hours": sum(shift["durationHours"] for shift in selected),
                    "shiftIds": [shift["id"] for shift in selected],
                }
            )
            start_index = None
    return blocks


def parse_schedule(workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    worksheet, table = find_contact_table(workbook)
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)

    headers = {
        col: str(worksheet.cell(min_row, col).value or "").strip()
        for col in range(min_col, max_col + 1)
    }
    name_columns = [column for column, header in headers.items() if header == "Név"]
    if len(name_columns) != 1:
        raise ValueError("A 'contact' táblában pontosan egy 'Név' oszlop szükséges.")
    name_column = name_columns[0]

    shift_columns: list[tuple[int, dict[str, Any]]] = []
    for column in range(min_col, max_col + 1):
        match = SHIFT_HEADER.fullmatch(headers[column])
        if not match:
            continue
        raw_day, start_hour, start_minute, end_hour, end_minute = match.groups()
        day = DAY_NAMES[raw_day.casefold()]
        start = f"{int(start_hour):02d}:{start_minute}"
        end = f"{int(end_hour):02d}:{end_minute}"
        shift_id = f"{slugify(day)}-{start.replace(':', '')}-{end.replace(':', '')}"
        shift_columns.append(
            (
                column,
                {
                    "id": shift_id,
                    "day": day,
                    "start": start,
                    "end": end,
                    "label": f"{day}\n{start} – {end}",
                    "durationHours": duration_hours(
                        int(start_hour), int(start_minute), int(end_hour), int(end_minute)
                    ),
                    "order": len(shift_columns),
                },
            )
        )
    if not shift_columns:
        raise ValueError("Nem található Sze/Cs/P/Szo kezdetű, időintervallumot tartalmazó turnusoszlop.")

    shifts = [shift for _, shift in shift_columns]
    workers: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for row in range(min_row + 1, max_row + 1):
        raw_name = worksheet.cell(row, name_column).value
        if raw_name is None or not str(raw_name).strip():
            continue
        assignments = {
            shift["id"]: is_x(worksheet.cell(row, column).value)
            for column, shift in shift_columns
        }
        if not any(assignments.values()):
            continue
        name = re.sub(r"\s+", " ", str(raw_name)).strip()
        identifier = worker_id(name)
        if identifier in seen_ids:
            raise ValueError(f"Duplikált, beosztott név található: {name}")
        seen_ids.add(identifier)
        worker_shifts = [shift for shift in shifts if assignments[shift["id"]]]
        workers.append(
            {
                "id": identifier,
                "name": name,
                "assignments": assignments,
                "shiftIds": [shift["id"] for shift in worker_shifts],
                "scheduledHours": sum(shift["durationHours"] for shift in worker_shifts),
                "blocks": continuous_blocks(assignments, shifts),
            }
        )
    workers.sort(key=lambda worker: sort_key(worker["name"]))

    boundaries: list[dict[str, Any]] = []
    for index, shift in enumerate(shifts):
        previous = shifts[index - 1] if index > 0 else None
        arrivals = []
        departures = []
        for worker in workers:
            current_active = worker["assignments"][shift["id"]]
            previous_active = bool(previous and worker["assignments"][previous["id"]])
            if current_active and not previous_active:
                arrivals.append(worker["id"])
            if previous_active and not current_active:
                departures.append(worker["id"])
        boundaries.append(
            {
                "id": f'{shift["id"]}-start',
                "label": f'{shift["day"]} {shift["start"]}',
                "currentShiftId": shift["id"],
                "previousShiftId": previous["id"] if previous else None,
                "arrivals": arrivals,
                "departures": departures,
                "order": len(boundaries),
            }
        )

    last_shift = shifts[-1]
    boundaries.append(
        {
            "id": "event-end",
            "label": f'{last_shift["day"]} {last_shift["end"]} – rendezvény vége',
            "currentShiftId": None,
            "previousShiftId": last_shift["id"],
            "arrivals": [],
            "departures": [
                worker["id"]
                for worker in workers
                if worker["assignments"][last_shift["id"]]
            ],
            "order": len(boundaries),
        }
    )

    return {
        "event": "Siroki Motoros találkozó 2026 - munkainformációk",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "source": {
            "workbook": workbook_path.name,
            "sheet": worksheet.title,
            "table": table.name,
        },
        "shifts": shifts,
        "workers": workers,
        "boundaries": boundaries,
        "statistics": {
            "workerCount": len(workers),
            "shiftCount": len(shifts),
            "scheduledHours": sum(worker["scheduledHours"] for worker in workers),
        },
    }


def render_markdown(source: str) -> str:
    if markdown is not None:
        return markdown.markdown(
            source,
            extensions=["extra", "sane_lists"],
            output_format="html5",
        )
    # Dependency-free fallback for local previews. GitHub Actions uses Python-Markdown.
    output: list[str] = []
    paragraph: list[str] = []
    in_list = False

    def inline(text: str) -> str:
        escaped = html.escape(text)
        escaped = re.sub(r"`([^`]+)`", r"<code>\1</code>", escaped)
        escaped = re.sub(r"\*\*([^*]+)\*\*", r"<strong>\1</strong>", escaped)
        return escaped

    def flush_paragraph() -> None:
        if paragraph:
            output.append(f'<p>{inline(" ".join(paragraph))}</p>')
            paragraph.clear()

    for raw_line in source.splitlines() + [""]:
        line = raw_line.strip()
        heading = re.match(r"^(#{1,6})\s+(.+)$", line)
        if heading:
            flush_paragraph()
            if in_list:
                output.append("</ul>")
                in_list = False
            level = len(heading.group(1))
            output.append(f"<h{level}>{inline(heading.group(2))}</h{level}>")
        elif line.startswith(("- ", "* ")):
            flush_paragraph()
            if not in_list:
                output.append("<ul>")
                in_list = True
            output.append(f"<li>{inline(line[2:])}</li>")
        elif not line:
            flush_paragraph()
            if in_list:
                output.append("</ul>")
                in_list = False
        else:
            paragraph.append(line)
    return "\n".join(output)


def empty_cars(schedule: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {
            "boundary_id": boundary["id"],
            "payload": {"arrivals": {"cars": []}, "departures": {"cars": []}},
        }
        for boundary in schedule["boundaries"]
    ]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=Path("data/Sirok 2026 beosztás.xlsx"))
    parser.add_argument("--markdown", type=Path, default=Path("data/munkainformaciok.md"))
    parser.add_argument("--output", type=Path, default=Path("site/data"))
    args = parser.parse_args()

    args.output.mkdir(parents=True, exist_ok=True)
    schedule = parse_schedule(args.workbook)
    (args.output / "schedule.json").write_text(
        json.dumps(schedule, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (args.output / "info.html").write_text(
        render_markdown(args.markdown.read_text(encoding="utf-8")), encoding="utf-8"
    )
    cars_path = args.output / "cars.json"
    if not cars_path.exists():
        cars_path.write_text(
            json.dumps(empty_cars(schedule), ensure_ascii=False, indent=2), encoding="utf-8"
        )
    print(
        f'Elkészült: {len(schedule["workers"])} dolgozó, '
        f'{len(schedule["shifts"])} turnus, {len(schedule["boundaries"])} váltási időpont.'
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
